#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
委托加工管理系统
Python + Flask + SQLite + 浏览器界面
"""

import os
import sys
import re
import io
import glob as gglob
import shutil
from datetime import datetime, timedelta
from decimal import Decimal
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    openpyxl = None
from flask import (Flask, render_template, request, redirect, url_for,
                   flash, jsonify, send_file)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, text, or_, and_
from sqlalchemy.orm import joinedload

# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------
# 检测是否在 PyInstaller 打包环境中运行
if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    BASE_DIR = sys._MEIPASS  # 打包后的临时解压目录（_internal）
    DATA_DIR = os.path.join(os.path.dirname(sys.executable), 'data')
    # 首次运行时从包内复制数据库到数据目录
    src_db = os.path.join(BASE_DIR, 'data', 'database.db')
    if not os.path.exists(DATA_DIR) and os.path.exists(src_db):
        os.makedirs(DATA_DIR, exist_ok=True)
        shutil.copy2(src_db, os.path.join(DATA_DIR, 'database.db'))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DATA_DIR = os.path.join(BASE_DIR, 'data')

DB_PATH = os.path.join(DATA_DIR, 'database.db')

app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, 'templates'))
app.secret_key = 'weituojiagong-2026-secret-key'
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}?charset=utf-8'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# 确保data目录存在
os.makedirs(DATA_DIR, exist_ok=True)

# 已知加工户名（Excel导入时从文件名/备注中提取加工户名称）
KNOWN_PROCESSOR_NAMES = ['徐正玺', '孙宇', '李春花', '李炳春', '董红国', '钱艳', '高明', '杨忠良']

# ---------------------------------------------------------------------------
# 模板上下文处理器
# ---------------------------------------------------------------------------
@app.context_processor
def utility_processor():
    return {'now': lambda: datetime.now().strftime('%Y-%m-%d %H:%M'),
            'today': lambda: datetime.now().strftime('%Y-%m-%d')}

# ---------------------------------------------------------------------------
# 数据库模型
# ---------------------------------------------------------------------------

class Processor(db.Model):
    """加工户"""
    __tablename__ = 'processors'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    phone = db.Column(db.String(50))
    address = db.Column(db.String(200))
    process_step = db.Column(db.String(50), default='')  # 工序：皮壳、服装、后道、充棉等
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)

    shipments = db.relationship('Shipment', backref='processor', lazy='dynamic')
    receipts = db.relationship('Receipt', backref='processor', lazy='dynamic')
    prices = db.relationship('ProcessorPrice', backref='processor', lazy='dynamic')

    def to_dict(self):
        return {'id': self.id, 'name': self.name, 'phone': self.phone or '',
                'address': self.address or '', 'process_step': self.process_step or '',
                'notes': self.notes or ''}


class ProductCategory(db.Model):
    """产品分类"""
    __tablename__ = 'product_categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False, unique=True)

    products = db.relationship('Product', backref='category', lazy='dynamic')


class Product(db.Model):
    """产品/半成品/原材料"""
    __tablename__ = 'products'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    unit = db.Column(db.String(20), default='个')  # 个/对/米/公斤/袋
    category_id = db.Column(db.Integer, db.ForeignKey('product_categories.id'))
    product_type = db.Column(db.String(20), default='成品')  # 成品/半成品/原材料
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)

    def to_dict(self):
        return {
            'id': self.id, 'name': self.name, 'unit': self.unit or '个',
            'category': self.category.name if self.category else '',
            'product_type': self.product_type or '成品',
            'notes': self.notes or ''
        }


class ProcessorPrice(db.Model):
    """加工单价 - 每个加工户对不同产品的单价"""
    __tablename__ = 'processor_prices'
    id = db.Column(db.Integer, primary_key=True)
    processor_id = db.Column(db.Integer, db.ForeignKey('processors.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    price = db.Column(db.Float, default=0)
    notes = db.Column(db.Text)

    product = db.relationship('Product', backref='prices')


class Shipment(db.Model):
    """发货单"""
    __tablename__ = 'shipments'
    id = db.Column(db.Integer, primary_key=True)
    processor_id = db.Column(db.Integer, db.ForeignKey('processors.id'), nullable=False)
    shipment_date = db.Column(db.String(20), nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)

    items = db.relationship('ShipmentItem', backref='shipment', lazy='dynamic',
                            cascade='all, delete-orphan')


class ShipmentItem(db.Model):
    """发货明细"""
    __tablename__ = 'shipment_items'
    id = db.Column(db.Integer, primary_key=True)
    shipment_id = db.Column(db.Integer, db.ForeignKey('shipments.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    quantity = db.Column(db.Float, default=0)
    spec = db.Column(db.String(100))  # 规格/颜色
    notes = db.Column(db.Text)

    product = db.relationship('Product', backref='shipment_items')


class Receipt(db.Model):
    """收货单"""
    __tablename__ = 'receipts'
    id = db.Column(db.Integer, primary_key=True)
    processor_id = db.Column(db.Integer, db.ForeignKey('processors.id'), nullable=False)
    receipt_date = db.Column(db.String(20), nullable=False)
    total_amount = db.Column(db.Float, default=0)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)

    items = db.relationship('ReceiptItem', backref='receipt', lazy='dynamic',
                            cascade='all, delete-orphan')


class ReceiptItem(db.Model):
    """收货明细"""
    __tablename__ = 'receipt_items'
    id = db.Column(db.Integer, primary_key=True)
    receipt_id = db.Column(db.Integer, db.ForeignKey('receipts.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    quantity = db.Column(db.Float, default=0)
    unit_price = db.Column(db.Float, default=0)
    amount = db.Column(db.Float, default=0)
    spec = db.Column(db.String(100))
    notes = db.Column(db.Text)

    product = db.relationship('Product', backref='receipt_items')


class TransferRecord(db.Model):
    """转单记录"""
    __tablename__ = 'transfer_records'
    id = db.Column(db.Integer, primary_key=True)
    from_processor_id = db.Column(db.Integer, db.ForeignKey('processors.id'), nullable=False)
    to_processor_id = db.Column(db.Integer, db.ForeignKey('processors.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'))
    quantity = db.Column(db.Float, default=0)
    price = db.Column(db.Float, nullable=True)  # 加工单价，为空时从 ProcessorPrice 取
    transfer_date = db.Column(db.String(20))
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)

    from_processor = db.relationship('Processor', foreign_keys=[from_processor_id])
    to_processor = db.relationship('Processor', foreign_keys=[to_processor_id])
    product = db.relationship('Product')


class BeginInventory(db.Model):
    """期初库存"""
    __tablename__ = 'begin_inventory'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    processor_id = db.Column(db.Integer, db.ForeignKey('processors.id'))
    quantity = db.Column(db.Float, default=0)
    period = db.Column(db.String(20), default='2026-02')  # 所属期

    product = db.relationship('Product')
    processor = db.relationship('Processor')


# ---------------------------------------------------------------------------
# 辅助函数
# ---------------------------------------------------------------------------

def get_inventory(processor_id=None, product_id=None, period=None):
    """计算收发存汇总（含转单）
    结存 = 期初 + 发货 + 转入 - 收货 - 转出
    注：发货+转入视为进入加工户库存，收货+转出视为减少加工户库存。
       转单数据自动纳入计算，正数表示货还在加工户手里。
   如果指定 period（日期字符串如 '2026-02-28'），则只统计该日期前的数据。
    """
    bi_filter = " AND bi.period = :period" if period else ""
    ship_filter = " AND sh.shipment_date <= :period" if period else ""
    receipt_filter = " AND rc.receipt_date <= :period" if period else ""
    transfer_filter = " AND t.transfer_date <= :period" if period else ""

    query = f"""
    SELECT
        p.id AS product_id,
        p.name AS product_name,
        p.unit,
        p.product_type,
        pr.id AS processor_id,
        pr.name AS processor_name,
        COALESCE(bi.quantity, 0) AS begin_qty,
        COALESCE(s.total_ship, 0) AS ship_qty,
        COALESCE(ti.total_in, 0) AS transfer_in_qty,
        COALESCE(r.total_receipt, 0) AS receipt_qty,
        COALESCE(tr_out.total_out, 0) AS transfer_out_qty,
        COALESCE(bi.quantity, 0) + COALESCE(s.total_ship, 0) + COALESCE(ti.total_in, 0)
        - COALESCE(r.total_receipt, 0) - COALESCE(tr_out.total_out, 0) AS balance
    FROM products p
    CROSS JOIN processors pr
    LEFT JOIN begin_inventory bi ON bi.product_id = p.id AND bi.processor_id = pr.id
        {bi_filter}
    LEFT JOIN (
        SELECT si.product_id, sh.processor_id, SUM(si.quantity) AS total_ship
        FROM shipment_items si
        JOIN shipments sh ON sh.id = si.shipment_id
        WHERE 1=1{ship_filter}
        GROUP BY si.product_id, sh.processor_id
    ) s ON s.product_id = p.id AND s.processor_id = pr.id
    LEFT JOIN (
        SELECT ri.product_id, rc.processor_id, SUM(ri.quantity) AS total_receipt
        FROM receipt_items ri
        JOIN receipts rc ON rc.id = ri.receipt_id
        WHERE 1=1{receipt_filter}
        GROUP BY ri.product_id, rc.processor_id
    ) r ON r.product_id = p.id AND r.processor_id = pr.id
    LEFT JOIN (
        SELECT to_processor_id AS processor_id, product_id, SUM(quantity) AS total_in
        FROM transfer_records t
        WHERE product_id IS NOT NULL{transfer_filter}
        GROUP BY to_processor_id, product_id
    ) ti ON ti.processor_id = pr.id AND ti.product_id = p.id
    LEFT JOIN (
        SELECT from_processor_id AS processor_id, product_id, SUM(quantity) AS total_out
        FROM transfer_records t
        WHERE product_id IS NOT NULL{transfer_filter}
        GROUP BY from_processor_id, product_id
    ) tr_out ON tr_out.processor_id = pr.id AND tr_out.product_id = p.id
    WHERE 1=1
        AND (COALESCE(bi.quantity, 0) != 0
             OR COALESCE(s.total_ship, 0) != 0
             OR COALESCE(r.total_receipt, 0) != 0
             OR COALESCE(ti.total_in, 0) != 0
             OR COALESCE(tr_out.total_out, 0) != 0)
    """
    params = {}
    if period:
        params['period'] = period
    if processor_id:
        query += " AND pr.id = :pid"
        params['pid'] = processor_id
    if product_id:
        query += " AND p.id = :prodid"
        params['prodid'] = product_id

    query += " ORDER BY pr.name, p.name"
    result = db.session.execute(text(query), params).fetchall()
    return result


# ---------------------------------------------------------------------------
# 路由 - 首页 / 仪表盘
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    stats = {
        'processors': Processor.query.count(),
        'products': Product.query.count(),
        'shipments': Shipment.query.count(),
        'receipts': Receipt.query.count(),
    }
    # 最近发货
    recent_shipments = Shipment.query.order_by(Shipment.created_at.desc()).limit(5).all()
    recent_receipts = Receipt.query.order_by(Receipt.created_at.desc()).limit(5).all()

    # 异常库存（结存为负数的）
    inv_data = get_inventory()
    anomalies = []
    for row in inv_data:
        if row.balance < -0.01:
            anomalies.append(row)

    return render_template('dashboard.html', stats=stats,
                           recent_shipments=recent_shipments,
                           recent_receipts=recent_receipts,
                           anomalies=anomalies[:10])


# ---------------------------------------------------------------------------
# 路由 - 加工户管理
# ---------------------------------------------------------------------------

@app.route('/processors')
def processor_list():
    processors = Processor.query.order_by(Processor.name).all()
    return render_template('processors.html', processors=processors)


@app.route('/processors/add', methods=['POST'])
def processor_add():
    name = request.form.get('name', '').strip()
    if not name:
        flash('请输入加工户名称', 'danger')
        return redirect(url_for('processor_list'))
    if Processor.query.filter_by(name=name).first():
        flash(f'加工户 "{name}" 已存在', 'warning')
        return redirect(url_for('processor_list'))
    p = Processor(name=name, phone=request.form.get('phone', ''),
                  address=request.form.get('address', ''),
                  process_step=request.form.get('process_step', ''),
                  notes=request.form.get('notes', ''))
    db.session.add(p)
    db.session.commit()
    flash(f'加工户 "{name}" 添加成功', 'success')
    return redirect(url_for('processor_list'))


@app.route('/processors/edit/<int:id>', methods=['POST'])
def processor_edit(id):
    p = Processor.query.get_or_404(id)
    name = request.form.get('name', '').strip()
    if not name:
        flash('请输入加工户名称', 'danger')
        return redirect(url_for('processor_list'))
    existing = Processor.query.filter_by(name=name).first()
    if existing and existing.id != id:
        flash(f'加工户 "{name}" 已存在', 'warning')
        return redirect(url_for('processor_list'))
    p.name = name
    p.phone = request.form.get('phone', '')
    p.address = request.form.get('address', '')
    p.process_step = request.form.get('process_step', '')
    p.notes = request.form.get('notes', '')
    db.session.commit()
    flash('修改成功', 'success')
    return redirect(url_for('processor_list'))


@app.route('/processors/delete/<int:id>', methods=['POST'])
def processor_delete(id):
    p = Processor.query.get_or_404(id)
    if p.shipments.count() > 0 or p.receipts.count() > 0:
        flash(f'加工户 "{p.name}" 下有业务单据，不能删除', 'danger')
        return redirect(url_for('processor_list'))
    db.session.delete(p)
    db.session.commit()
    flash(f'已删除加工户 "{p.name}"', 'success')
    return redirect(url_for('processor_list'))


# ---------------------------------------------------------------------------
# 路由 - 产品管理
# ---------------------------------------------------------------------------

@app.route('/products')
def product_list():
    # 产品和单价合并管理，跳转到单价页面
    return redirect(url_for('price_list'))


@app.route('/products/add', methods=['POST'])
def product_add():
    name = request.form.get('name', '').strip()
    processor_id = request.form.get('processor_id', type=int)
    if not name:
        flash('请输入产品名称', 'danger')
        return redirect(url_for('price_list', processor_id=processor_id))
    p = Product(name=name, unit=request.form.get('unit', '个'),
                product_type=request.form.get('product_type', '成品'),
                category_id=request.form.get('category_id') or None,
                notes=request.form.get('notes', ''))
    db.session.add(p)
    db.session.commit()
    flash(f'产品 "{name}" 添加成功', 'success')
    return redirect(url_for('price_list', processor_id=processor_id))


@app.route('/products/edit/<int:id>', methods=['POST'])
def product_edit(id):
    p = Product.query.get_or_404(id)
    name = request.form.get('name', '').strip()
    if not name:
        flash('请输入产品名称', 'danger')
        return redirect(url_for('product_list'))
    p.name = name
    p.unit = request.form.get('unit', '个')
    p.product_type = request.form.get('product_type', '成品')
    p.category_id = request.form.get('category_id') or None
    p.notes = request.form.get('notes', '')
    db.session.commit()
    flash('修改成功', 'success')
    return redirect(url_for('product_list'))


@app.route('/products/delete/<int:id>', methods=['POST'])
def product_delete(id):
    p = Product.query.get_or_404(id)
    used_ship = ShipmentItem.query.filter_by(product_id=id).first()
    used_receipt = ReceiptItem.query.filter_by(product_id=id).first()
    if used_ship or used_receipt:
        flash(f'产品 "{p.name}" 已被业务单据引用，不能删除', 'danger')
        return redirect(url_for('product_list'))
    db.session.delete(p)
    db.session.commit()
    flash(f'已删除产品 "{p.name}"', 'success')
    return redirect(url_for('product_list'))


# ---------------------------------------------------------------------------
# 路由 - 加工单价管理
# ---------------------------------------------------------------------------

@app.route('/prices')
def price_list():
    processors = Processor.query.order_by(Processor.name).all()
    all_products = Product.query.order_by(Product.name).all()
    pid = request.args.get('processor_id', type=int)

    # 每个加工户的产品数量（批量查询避免 N+1）
    counts = db.session.query(
        ProcessorPrice.processor_id, func.count(ProcessorPrice.id)
    ).group_by(ProcessorPrice.processor_id).all()
    processor_product_counts = {pid: cnt for pid, cnt in counts}

    # 当前选中加工户的产品价格列表
    processor_prices = []
    current_processor = None
    if pid:
        current_processor = Processor.query.get(pid)
        if current_processor:
            processor_prices = db.session.query(ProcessorPrice).join(Product).filter(
                ProcessorPrice.processor_id == pid
            ).order_by(Product.name).all()

    return render_template('prices.html', processors=processors,
                           all_products=all_products,
                           processor_prices=processor_prices,
                           processor_product_counts=processor_product_counts,
                           current_pid=pid,
                           current_processor=current_processor)


@app.route('/prices/save', methods=['POST'])
def price_save():
    processor_id = request.form.get('processor_id', type=int)
    product_id = request.form.get('product_id', type=int)
    price_val = request.form.get('price', type=float, default=0)

    if not processor_id or not product_id:
        flash('请选择加工户和产品', 'danger')
        return redirect(url_for('price_list'))

    existing = ProcessorPrice.query.filter_by(
        processor_id=processor_id, product_id=product_id).first()
    if existing:
        existing.price = price_val
    else:
        pp = ProcessorPrice(processor_id=processor_id,
                            product_id=product_id, price=price_val)
        db.session.add(pp)
    db.session.flush()

    # 同步更新转单记录中该加工户+该产品的单价
    TransferRecord.query.filter_by(
        from_processor_id=processor_id, product_id=product_id
    ).update({TransferRecord.price: price_val}, synchronize_session=False)

    db.session.commit()
    flash('单价保存成功', 'success')
    return redirect(url_for('price_list', processor_id=processor_id))


@app.route('/prices/delete/<int:id>', methods=['POST'])
def price_delete(id):
    pp = ProcessorPrice.query.get_or_404(id)
    db.session.delete(pp)
    db.session.commit()
    flash('单价已删除', 'success')
    return redirect(url_for('price_list', processor_id=pp.processor_id))


@app.route('/api/prices/<int:processor_id>/<int:product_id>')
def api_price(processor_id, product_id):
    pp = ProcessorPrice.query.filter_by(
        processor_id=processor_id, product_id=product_id).first()
    return jsonify({'price': pp.price if pp else 0})


@app.route('/api/prices/save', methods=['POST'])
def api_price_save():
    """行内编辑保存单价"""
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'msg': '无效请求'})
    processor_id = data.get('processor_id')
    product_id = data.get('product_id')
    price = data.get('price', 0)
    if not processor_id or not product_id:
        return jsonify({'ok': False, 'msg': '参数不全'})
    pp = ProcessorPrice.query.filter_by(
        processor_id=processor_id, product_id=product_id).first()
    if pp:
        pp.price = price
    else:
        pp = ProcessorPrice(processor_id=processor_id,
                            product_id=product_id, price=price)
        db.session.add(pp)
    db.session.flush()

    # 同步更新转单记录中该加工户+该产品的单价
    TransferRecord.query.filter_by(
        from_processor_id=processor_id, product_id=product_id
    ).update({TransferRecord.price: price}, synchronize_session=False)

    db.session.commit()
    return jsonify({'ok': True, 'id': pp.id, 'price': price})


# ---------------------------------------------------------------------------
# 路由 - 发货单管理
# ---------------------------------------------------------------------------

@app.route('/shipments')
def shipment_list():
    processor_id = request.args.get('processor_id', type=int)
    sort = request.args.get('sort', 'desc')
    query = Shipment.query
    if processor_id:
        query = query.filter_by(processor_id=processor_id)
    if sort == 'asc':
        shipments = query.order_by(Shipment.shipment_date.asc(),
                                   Shipment.id.asc()).all()
    else:
        shipments = query.order_by(Shipment.shipment_date.desc(),
                                   Shipment.id.desc()).all()
    processors = Processor.query.order_by(Processor.name).all()

    # 按(日期, 加工户)分组合并
    # 批量加载所有明细避免 N+1
    shipment_ids = [s.id for s in shipments]
    items_by_shipment = defaultdict(list)
    if shipment_ids:
        all_items = ShipmentItem.query.filter(
            ShipmentItem.shipment_id.in_(shipment_ids)
        ).options(joinedload(ShipmentItem.product)).all()
        for item in all_items:
            items_by_shipment[item.shipment_id].append(item)

    groups = defaultdict(list)
    for s in shipments:
        key = (s.shipment_date, s.processor_id, s.processor.name)
        products = []
        for item in items_by_shipment.get(s.id, []):
            products.append({
                'name': item.product.name,
                'quantity': item.quantity,
                'spec': item.spec or ''
            })
        notes = s.notes or ''
        groups[key].append({'products': products, 'notes': notes, 'id': s.id})

    grouped = []
    for (date, pid, pname), records in sorted(groups.items(), reverse=True):
        all_products = []
        all_notes = set()
        for r in records:
            all_products.extend(r['products'])
            if r['notes']:
                all_notes.add(r['notes'])
        grouped.append({
            'date': date,
            'processor_id': pid,
            'processor_name': pname,
            'products': all_products,
            'notes': '; '.join(all_notes),
            'ids': [r['id'] for r in records]
        })

    return render_template('shipments.html', shipments=grouped,
                           processors=processors, processor_id=processor_id,
                           sort=sort)


@app.route('/shipments/add', methods=['GET', 'POST'])
@app.route('/shipments/edit/<int:id>', methods=['GET', 'POST'])
def shipment_add(id=None):
    if request.method == 'GET':
        processors = Processor.query.order_by(Processor.name).all()
        products = Product.query.order_by(Product.name).all()
        shipment = None
        items = []
        if id:
            shipment = Shipment.query.get_or_404(id)
            items = ShipmentItem.query.filter_by(shipment_id=id).all()
        return render_template('shipment_form.html', processors=processors,
                               products=products, shipment=shipment, edit_items=items)

    # POST
    processor_id = request.form.get('processor_id', type=int)
    shipment_date = request.form.get('shipment_date', '')
    notes = request.form.get('notes', '')

    if not processor_id or not shipment_date:
        flash('请选择加工户和填写日期', 'danger')
        return redirect(url_for('shipment_add'))

    if id:
        shipment = Shipment.query.get_or_404(id)
        shipment.processor_id = processor_id
        shipment.shipment_date = shipment_date
        shipment.notes = notes
        ShipmentItem.query.filter_by(shipment_id=id).delete()
    else:
        shipment = Shipment(processor_id=processor_id,
                            shipment_date=shipment_date, notes=notes)
        db.session.add(shipment)
    db.session.flush()

    product_ids = request.form.getlist('product_id[]')
    quantities = request.form.getlist('quantity[]')
    specs = request.form.getlist('spec[]')
    item_notes = request.form.getlist('item_notes[]')

    for i in range(len(product_ids)):
        pid = int(product_ids[i]) if product_ids[i] else None
        qty = float(quantities[i]) if quantities[i] else 0
        if pid and qty != 0:
            item = ShipmentItem(
                shipment_id=shipment.id,
                product_id=pid,
                quantity=qty,
                spec=specs[i] if i < len(specs) else '',
                notes=item_notes[i] if i < len(item_notes) else ''
            )
            db.session.add(item)

    db.session.commit()
    flash('发货单保存成功', 'success')
    return redirect(url_for('shipment_list'))


@app.route('/shipments/view/<int:id>')
def shipment_view(id):
    shipment = Shipment.query.get_or_404(id)
    items = ShipmentItem.query.filter_by(shipment_id=id).all()
    return render_template('shipment_detail.html', shipment=shipment, items=items)


@app.route('/shipments/delete/<int:id>', methods=['POST'])
def shipment_delete(id):
    s = Shipment.query.get_or_404(id)
    db.session.delete(s)
    db.session.commit()
    flash('发货单已删除', 'success')
    return redirect(url_for('shipment_list'))


# ---------------------------------------------------------------------------
# 路由 - 收货单管理
# ---------------------------------------------------------------------------

@app.route('/receipts')
def receipt_list():
    processor_id = request.args.get('processor_id', type=int)
    sort = request.args.get('sort', 'desc')
    query = Receipt.query
    if processor_id:
        query = query.filter_by(processor_id=processor_id)
    if sort == 'asc':
        receipts = query.order_by(Receipt.receipt_date.asc(),
                                  Receipt.id.asc()).all()
    else:
        receipts = query.order_by(Receipt.receipt_date.desc(),
                                  Receipt.id.desc()).all()
    processors = Processor.query.order_by(Processor.name).all()

    # 按(日期, 加工户)分组合并
    # 批量加载所有明细避免 N+1
    receipt_ids = [r.id for r in receipts]
    items_by_receipt = defaultdict(list)
    if receipt_ids:
        all_items = ReceiptItem.query.filter(
            ReceiptItem.receipt_id.in_(receipt_ids)
        ).options(joinedload(ReceiptItem.product)).all()
        for item in all_items:
            items_by_receipt[item.receipt_id].append(item)

    groups = defaultdict(list)
    for r in receipts:
        key = (r.receipt_date, r.processor_id, r.processor.name)
        products = []
        for item in items_by_receipt.get(r.id, []):
            products.append({
                'name': item.product.name,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'amount': item.amount,
                'spec': item.spec or ''
            })
        groups[key].append({
            'products': products, 'notes': r.notes or '',
            'total_amount': r.total_amount or 0, 'id': r.id
        })

    grouped = []
    for (date, pid, pname), records in sorted(groups.items(), reverse=True):
        all_products = []
        total = 0
        all_notes = set()
        for rec in records:
            all_products.extend(rec['products'])
            total += rec['total_amount']
            if rec['notes']:
                all_notes.add(rec['notes'])
        grouped.append({
            'date': date,
            'processor_id': pid,
            'processor_name': pname,
            'products': all_products,
            'total_amount': total,
            'notes': '; '.join(all_notes),
            'ids': [rec['id'] for rec in records]
        })

    return render_template('receipts.html', receipts=grouped,
                           processors=processors, processor_id=processor_id,
                           sort=sort)


@app.route('/receipts/add', methods=['GET', 'POST'])
def receipt_add():
    if request.method == 'GET':
        processors = Processor.query.order_by(Processor.name).all()
        products = Product.query.order_by(Product.name).all()
        return render_template('receipt_form.html', processors=processors,
                               products=products, receipt=None)

    processor_id = request.form.get('processor_id', type=int)
    receipt_date = request.form.get('receipt_date', '')
    notes = request.form.get('notes', '')
    product_ids = request.form.getlist('product_id[]')
    quantities = request.form.getlist('quantity[]')
    unit_prices = request.form.getlist('unit_price[]')
    specs = request.form.getlist('spec[]')
    item_notes = request.form.getlist('item_notes[]')

    if not processor_id or not receipt_date:
        flash('请选择加工户和填写日期', 'danger')
        return redirect(url_for('receipt_add'))

    receipt = Receipt(processor_id=processor_id,
                      receipt_date=receipt_date, notes=notes)
    db.session.add(receipt)
    db.session.flush()

    total_amount = 0
    for i in range(len(product_ids)):
        pid = int(product_ids[i]) if product_ids[i] else None
        qty = float(quantities[i]) if quantities[i] else 0
        up = float(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else 0
        amt = qty * up
        if pid and qty != 0:
            item = ReceiptItem(
                receipt_id=receipt.id,
                product_id=pid,
                quantity=qty,
                unit_price=up,
                amount=amt,
                spec=specs[i] if i < len(specs) else '',
                notes=item_notes[i] if i < len(item_notes) else ''
            )
            db.session.add(item)
            total_amount += amt

    receipt.total_amount = total_amount
    db.session.commit()
    flash('收货单创建成功', 'success')
    return redirect(url_for('receipt_list'))


@app.route('/receipts/edit/<int:id>', methods=['GET', 'POST'])
def receipt_edit(id):
    receipt = Receipt.query.get_or_404(id)
    if request.method == 'GET':
        processors = Processor.query.order_by(Processor.name).all()
        products = Product.query.order_by(Product.name).all()
        items = ReceiptItem.query.filter_by(receipt_id=id).all()
        return render_template('receipt_form.html', processors=processors,
                               products=products, receipt=receipt, edit_items=items)

    # POST - update
    processor_id = request.form.get('processor_id', type=int)
    receipt_date = request.form.get('receipt_date', '')
    notes = request.form.get('notes', '')

    if not processor_id or not receipt_date:
        flash('请选择加工户和填写日期', 'danger')
        return redirect(url_for('receipt_edit', id=id))

    receipt.processor_id = processor_id
    receipt.receipt_date = receipt_date
    receipt.notes = notes
    ReceiptItem.query.filter_by(receipt_id=id).delete()
    db.session.flush()

    product_ids = request.form.getlist('product_id[]')
    quantities = request.form.getlist('quantity[]')
    unit_prices = request.form.getlist('unit_price[]')
    specs = request.form.getlist('spec[]')
    item_notes = request.form.getlist('item_notes[]')

    total_amount = 0
    for i in range(len(product_ids)):
        pid = int(product_ids[i]) if product_ids[i] else None
        qty = float(quantities[i]) if quantities[i] else 0
        up = float(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else 0
        amt = qty * up
        if pid and qty != 0:
            item = ReceiptItem(
                receipt_id=receipt.id,
                product_id=pid,
                quantity=qty,
                unit_price=up,
                amount=amt,
                spec=specs[i] if i < len(specs) else '',
                notes=item_notes[i] if i < len(item_notes) else ''
            )
            db.session.add(item)
            total_amount += amt

    receipt.total_amount = total_amount
    db.session.commit()
    flash('收货单已更新', 'success')
    return redirect(url_for('receipt_list'))


@app.route('/receipts/view/<int:id>')
def receipt_view(id):
    receipt = Receipt.query.get_or_404(id)
    items = ReceiptItem.query.filter_by(receipt_id=id).all()
    return render_template('receipt_detail.html', receipt=receipt, items=items)


@app.route('/receipts/delete/<int:id>', methods=['POST'])
def receipt_delete(id):
    r = Receipt.query.get_or_404(id)
    db.session.delete(r)
    db.session.commit()
    flash('收货单已删除', 'success')
    return redirect(url_for('receipt_list'))


# ---------------------------------------------------------------------------
# 路由 - 库存报表
# ---------------------------------------------------------------------------

@app.route('/inventory')
def inventory_report():
    processor_id = request.args.get('processor_id', type=int)
    anomaly_only = request.args.get('anomaly_only', type=int)

    inv_data = get_inventory(processor_id=processor_id)
    processors = Processor.query.order_by(Processor.name).all()

    results = []
    for row in inv_data:
        if anomaly_only and row.balance >= -0.01:
            continue
        results.append(row)

    return render_template('inventory.html', inventory=results,
                           processors=processors, processor_id=processor_id,
                           anomaly_only=anomaly_only)


# ---------------------------------------------------------------------------
# 路由 - 期初库存管理
# ---------------------------------------------------------------------------

@app.route('/begin_inventory')
def begin_inventory_list():
    processor_id = request.args.get('processor_id', type=int)
    period = request.args.get('period', '')
    processors = Processor.query.order_by(Processor.name).all()

    # 可用账期列表
    period_rows = db.session.query(BeginInventory.period).distinct().order_by(BeginInventory.period.desc()).all()
    available_periods = [r.period for r in period_rows]

    query = BeginInventory.query.options(
        joinedload(BeginInventory.product),
        joinedload(BeginInventory.processor)
    )
    if processor_id:
        query = query.filter(BeginInventory.processor_id == processor_id)
    if period:
        query = query.filter(BeginInventory.period == period)
    records = query.order_by(BeginInventory.processor_id, BeginInventory.product_id).all()

    all_products = Product.query.order_by(Product.name).all()

    return render_template('begin_inventory.html', records=records,
                           processors=processors, processor_id=processor_id,
                           all_products=all_products,
                           available_periods=available_periods, current_period=period)


@app.route('/begin_inventory/add', methods=['POST'])
def begin_inventory_add():
    processor_id = request.form.get('processor_id', type=int)
    product_id = request.form.get('product_id', type=int)
    quantity = request.form.get('quantity', type=float, default=0)
    period = request.form.get('period', '2026-02')

    if not processor_id or not product_id:
        flash('请选择加工户和产品', 'danger')
        return redirect(url_for('begin_inventory_list'))

    existing = BeginInventory.query.filter_by(
        processor_id=processor_id, product_id=product_id).first()
    if existing:
        existing.quantity = quantity
        existing.period = period
    else:
        bi = BeginInventory(processor_id=processor_id, product_id=product_id,
                            quantity=quantity, period=period)
        db.session.add(bi)
    db.session.commit()
    flash('期初库存已保存', 'success')
    return redirect(url_for('begin_inventory_list'))


@app.route('/begin_inventory/edit/<int:id>', methods=['POST'])
def begin_inventory_edit(id):
    bi = BeginInventory.query.get_or_404(id)
    quantity = request.form.get('quantity', type=float, default=0)
    bi.quantity = quantity
    db.session.commit()
    flash('期初库存已更新', 'success')
    return redirect(url_for('begin_inventory_list'))


@app.route('/begin_inventory/delete/<int:id>', methods=['POST'])
def begin_inventory_delete(id):
    bi = BeginInventory.query.get_or_404(id)
    db.session.delete(bi)
    db.session.commit()
    flash('期初库存已删除', 'success')
    return redirect(url_for('begin_inventory_list'))


# ---------------------------------------------------------------------------
# 路由 - 结账管理
# ---------------------------------------------------------------------------

def next_period(period):
    """计算下一个账期日期：2026-02-28 → 2026-03-01"""
    try:
        d = datetime.strptime(period, '%Y-%m-%d').date()
        return (d + timedelta(days=1)).strftime('%Y-%m-%d')
    except ValueError:
        return period


@app.route('/closing')
def closing_page():
    processors = Processor.query.order_by(Processor.name).all()

    # 可用账期列表
    period_rows = db.session.query(BeginInventory.period).distinct().order_by(BeginInventory.period.desc()).all()
    available_periods = [r.period for r in period_rows]

    today = datetime.now().strftime('%Y-%m-%d')

    # 默认：从最近一次结账日的次日起，到今天
    default_from = ''
    if available_periods:
        # 取最近一次结账后起始日（期初库存的所属期）
        default_from = available_periods[0]
    if not default_from:
        default_from = today

    from_date = request.args.get('from_date', default_from)
    to_date = request.args.get('to_date', today)

    next_period_str = next_period(to_date)

    # 检查是否已结过账到下一期
    has_next = BeginInventory.query.filter_by(period=next_period_str).count() > 0

    # 获取各加工户加工汇总（按收货计算应付金额，日期范围）
    receipt_where = " WHERE 1=1"
    receipt_params = {}
    if from_date:
        receipt_where += " AND rc.receipt_date >= :from_date"
        receipt_params['from_date'] = from_date
    if to_date:
        receipt_where += " AND rc.receipt_date <= :to_date"
        receipt_params['to_date'] = to_date
    receipt_summary_sql = f"""
    SELECT
        rc.processor_id,
        pr.name AS processor_name,
        ri.product_id,
        p.name AS product_name,
        p.unit,
        SUM(ri.quantity) AS total_qty,
        SUM(ri.amount) AS total_amount
    FROM receipt_items ri
    JOIN receipts rc ON rc.id = ri.receipt_id
    JOIN processors pr ON pr.id = rc.processor_id
    JOIN products p ON p.id = ri.product_id
    {receipt_where}
    GROUP BY rc.processor_id, ri.product_id
    ORDER BY pr.name, p.name
    """
    receipt_rows = db.session.execute(text(receipt_summary_sql), receipt_params).fetchall() if not has_next else []

    # 汇总各加工户应付总额
    processor_totals = {}
    for row in receipt_rows:
        pid = row.processor_id
        if pid not in processor_totals:
            processor_totals[pid] = {
                'processor_name': row.processor_name,
                'total_amount': 0,
            }
        processor_totals[pid]['total_amount'] += row.total_amount or 0

    return render_template('closing.html', processors=processors,
                           current_period=to_date,
                           next_period=next_period_str,
                           has_next=has_next,
                           receipt_rows=receipt_rows,
                           processor_totals=processor_totals,
                           available_periods=available_periods,
                           from_date=from_date, to_date=to_date)


@app.route('/closing/download', methods=['POST'])
def closing_download():
    """下载当期各加工户对账单 Excel"""
    if openpyxl is None:
        flash('缺少 openpyxl 库，无法生成 Excel', 'danger')
        return redirect(url_for('closing_page'))

    current_period = request.form.get('to_date', datetime.now().strftime('%Y-%m-%d'))

    inv_data = get_inventory(period=current_period)
    processors = Processor.query.order_by(Processor.name).all()

    wb = openpyxl.Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    # 按加工户分sheet
    grouped = defaultdict(list)
    for row in inv_data:
        grouped[row.processor_id].append(row)

    for pr in processors:
        rows = grouped.get(pr.id, [])
        ws = wb.create_sheet(title=pr.name[:31])  # sheet名最多31字符

        # 标题
        ws.merge_cells('A1:I1')
        ws['A1'] = f'{pr.name} — 委托加工对账单（{current_period}）'
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)

        # 表头
        headers = ['产品', '单位', '类型', '期初数量', '发货', '转入', '收货', '转出', '结存']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        # 数据行
        for i, row in enumerate(rows, 4):
            ws.cell(row=i, column=1, value=row.product_name)
            ws.cell(row=i, column=2, value=row.unit or '个')
            ws.cell(row=i, column=3, value=row.product_type)
            ws.cell(row=i, column=4, value=float(row.begin_qty or 0))
            ws.cell(row=i, column=5, value=float(row.ship_qty or 0))
            ws.cell(row=i, column=6, value=float(row.transfer_in_qty or 0))
            ws.cell(row=i, column=7, value=float(row.receipt_qty or 0))
            ws.cell(row=i, column=8, value=float(row.transfer_out_qty or 0))
            ws.cell(row=i, column=9, value=float(row.balance or 0))

        # 设置列宽
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True,
                     download_name=f'委托加工对账单_{current_period}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/closing/do', methods=['POST'])
def closing_do():
    """执行结账：当前结存 → 下一期期初库存"""
    current_period = request.form.get('to_date', datetime.now().strftime('%Y-%m-%d'))
    next_period_str = next_period(current_period)

    # 检查是否已结过
    if BeginInventory.query.filter_by(period=next_period_str).count() > 0:
        flash(f'{next_period_str} 已有期初数据，请先清理再结账', 'warning')
        return redirect(url_for('closing_page'))

    inv_data = get_inventory(period=current_period)
    created = 0
    for row in inv_data:
        if row.balance is None or abs(row.balance) < 0.001:
            continue
        bi = BeginInventory(
            product_id=row.product_id,
            processor_id=row.processor_id,
            quantity=row.balance,
            period=next_period_str,
        )
        db.session.add(bi)
        created += 1

    db.session.commit()
    flash(f'结账完成！已生成 {created} 条 {next_period_str} 期初库存记录', 'success')
    return redirect(url_for('begin_inventory_list', processor_id=''))


# ---------------------------------------------------------------------------
# 路由 - 财务对账
# ---------------------------------------------------------------------------

@app.route('/finance')
def finance():
    processor_id = request.args.get('processor_id', type=int)
    processors = Processor.query.order_by(Processor.name).all()

    params = {}
    summary_query = """
    SELECT
        rc.processor_id,
        pr.name AS processor_name,
        COUNT(rc.id) AS receipt_count,
        SUM(rc.total_amount) AS total_amount
    FROM receipts rc
    JOIN processors pr ON pr.id = rc.processor_id
    """
    if processor_id:
        summary_query += " WHERE rc.processor_id = :pid"
        params['pid'] = processor_id
    summary_query += " GROUP BY rc.processor_id ORDER BY pr.name"

    summary = db.session.execute(
        text(summary_query), params).fetchall()

    # 转单加工费汇总（转出方应收的加工费）
    transfer_fee_query = """
    SELECT
        pr.id,
        pr.name,
        COUNT(t.id) AS transfer_count,
        COALESCE(SUM(t.quantity * pp.price), 0) AS total_fee
    FROM processors pr
    JOIN transfer_records t ON t.from_processor_id = pr.id
    LEFT JOIN processor_prices pp ON pp.processor_id = t.from_processor_id AND pp.product_id = t.product_id
    """
    if processor_id:
        transfer_fee_query += " WHERE pr.id = :tf_pid"
        params['tf_pid'] = processor_id
    transfer_fee_query += " GROUP BY pr.id, pr.name ORDER BY pr.name"

    transfer_fees = db.session.execute(
        text(transfer_fee_query), params).fetchall()

    return render_template('finance.html', processors=processors,
                           summary=summary, processor_id=processor_id,
                           transfer_fees=transfer_fees)


# ---------------------------------------------------------------------------
# 路由 - 转单记录
# ---------------------------------------------------------------------------

@app.route('/transfers')
def transfer_list():
    processor_id = request.args.get('processor_id', type=int)
    processors = Processor.query.order_by(Processor.name).all()
    products = Product.query.order_by(Product.name).all()

    query = TransferRecord.query
    if processor_id:
        query = query.filter(
            or_(
                TransferRecord.from_processor_id == processor_id,
                TransferRecord.to_processor_id == processor_id
            )
        )
    transfers = query.order_by(TransferRecord.created_at.desc()).all()

    # 为每笔转单计算加工费 = 转出方单价 × 数量
    # 批量查询单价避免 N+1（仅对没有单独定价的转单查 ProcessorPrice）
    price_pairs = {(t.from_processor_id, t.product_id) for t in transfers
                   if t.product_id and t.from_processor_id and t.price is None}
    price_map = {}
    if price_pairs:
        clauses = [and_(ProcessorPrice.processor_id == pid, ProcessorPrice.product_id == prod_id)
                   for pid, prod_id in price_pairs]
        for pp in ProcessorPrice.query.filter(or_(*clauses)).all():
            price_map[(pp.processor_id, pp.product_id)] = pp.price or 0

    transfer_data = []
    total_fee = 0
    for t in transfers:
        # 优先使用转单上保存的单价，其次从 ProcessorPrice 取
        unit_price = t.price if t.price is not None else price_map.get((t.from_processor_id, t.product_id), 0)
        fee = unit_price * (t.quantity or 0)
        total_fee += fee
        transfer_data.append({
            'transfer': t,
            'unit_price': unit_price,
            'fee': fee,
        })

    return render_template('transfers.html', processors=processors,
                           products=products, transfers=transfer_data,
                           total_fee=total_fee, processor_id=processor_id)


@app.route('/transfers/add', methods=['POST'])
def transfer_add():
    from_pid = request.form.get('from_processor_id', type=int)
    to_pid = request.form.get('to_processor_id', type=int)
    transfer_date = request.form.get('transfer_date', '')
    notes = request.form.get('notes', '')

    if not from_pid or not to_pid:
        flash('请选择转出和转入加工户', 'danger')
        return redirect(url_for('transfer_list'))

    product_ids = request.form.getlist('product_id[]')
    quantities = request.form.getlist('quantity[]')
    item_notes = request.form.getlist('item_notes[]')

    created = 0
    for i in range(len(product_ids)):
        pid = int(product_ids[i]) if product_ids[i] else None
        qty = float(quantities[i]) if quantities[i] else 0
        if pid and qty != 0:
            t = TransferRecord(
                from_processor_id=from_pid,
                to_processor_id=to_pid,
                product_id=pid,
                quantity=qty,
                transfer_date=transfer_date,
                notes=item_notes[i] if i < len(item_notes) else ''
            )
            db.session.add(t)
            created += 1

    if created == 0:
        flash('请至少添加一条产品记录', 'danger')
        return redirect(url_for('transfer_list'))

    db.session.commit()
    flash(f'转单记录已添加（{created} 条）', 'success')
    return redirect(url_for('transfer_list'))


@app.route('/api/transfers/<int:id>')
def api_transfer_get(id):
    """返回单条转单记录 JSON"""
    t = TransferRecord.query.get_or_404(id)
    return jsonify({
        'id': t.id,
        'from_processor_id': t.from_processor_id,
        'to_processor_id': t.to_processor_id,
        'product_id': t.product_id,
        'quantity': t.quantity,
        'price': t.price,
        'transfer_date': t.transfer_date or '',
        'notes': t.notes or '',
    })


@app.route('/transfers/edit/<int:id>', methods=['POST'])
def transfer_edit(id):
    t = TransferRecord.query.get_or_404(id)
    from_pid = request.form.get('from_processor_id', type=int)
    to_pid = request.form.get('to_processor_id', type=int)
    product_id = request.form.get('product_id', type=int)
    quantity = request.form.get('quantity', type=float, default=0)
    price = request.form.get('price', type=float)
    transfer_date = request.form.get('transfer_date', '')
    notes = request.form.get('notes', '')

    if not from_pid or not to_pid or not product_id or quantity == 0:
        flash('请完整填写转单信息', 'danger')
        return redirect(url_for('transfer_list'))

    t.from_processor_id = from_pid
    t.to_processor_id = to_pid
    t.product_id = product_id
    t.quantity = quantity
    t.price = price
    t.transfer_date = transfer_date
    t.notes = notes
    db.session.commit()
    flash('转单记录已更新', 'success')
    return redirect(url_for('transfer_list'))


@app.route('/transfers/delete/<int:id>', methods=['POST'])
def transfer_delete(id):
    t = TransferRecord.query.get_or_404(id)
    db.session.delete(t)
    db.session.commit()
    flash('转单记录已删除', 'success')
    return redirect(url_for('transfer_list'))


# ---------------------------------------------------------------------------
# 路由 - 数据导入（从Excel）
# ---------------------------------------------------------------------------

def extract_processor_name(fname):
    """从文件名中提取加工户名称"""
    name = fname.replace('.xlsx', '')
    for kn in KNOWN_PROCESSOR_NAMES:
        if kn in name:
            return kn
    # 尝试取最后一个中文人名（2-3个字）
    matches = re.findall(r'[一-鿿]{2,3}', name)
    if matches:
        return matches[-1]
    return name


def extract_transfer_target(remark):
    """从备注中提取转单目标加工户，如 '转钱艳' → '钱艳'"""
    for kn in KNOWN_PROCESSOR_NAMES:
        if kn in remark:
            return kn
    # 尝试取"转"后面的中文人名
    m = re.search(r'转[写]?\s*([一-鿿]{2,3})', remark)
    if m:
        return m.group(1)
    return None


@app.route('/import')
def import_page():
    return render_template('import.html')


@app.route('/import/do', methods=['POST'])
def import_do():
    """从压缩包目录导入Excel数据"""
    zip_dir = request.form.get('source_dir', '').strip()
    if not zip_dir or not os.path.isdir(zip_dir):
        # 尝试默认路径
        default_dir = os.path.join(
            BASE_DIR, '..',
            'xwechat_files', 'wxid_ufogibn91jth22_59b4',
            'msg', 'file', '2026-05', 'extracted'
        )
        if os.path.isdir(default_dir):
            zip_dir = default_dir
        else:
            flash(f'找不到数据目录: {zip_dir}', 'danger')
            return redirect(url_for('import_page'))

    xlsx_files = sorted(gglob.glob(os.path.join(zip_dir, '*.xlsx')))
    if not xlsx_files:
        flash(f'目录中没有Excel文件: {zip_dir}', 'warning')
        return redirect(url_for('import_page'))

    report = {'files': [], 'total_products': 0, 'total_shipments': 0,
              'total_receipts': 0, 'warnings': [], 'errors': []}

    # 去重：同一笔转单只录一次（内存+数据库双重检查）
    transfer_seen = set()
    def add_transfer(from_pid, to_pid, prod_id, qty, transfer_date=None):
        key = (from_pid, to_pid, prod_id, qty)
        if key in transfer_seen:
            return False
        existing = TransferRecord.query.filter_by(
            from_processor_id=from_pid, to_processor_id=to_pid,
            product_id=prod_id, quantity=qty).first()
        if existing:
            transfer_seen.add(key)
            return False
        # 检查是否已有同一天同加工户同数量的转单（产品不同视为同一笔）
        if transfer_date:
            existing2 = TransferRecord.query.filter(
                TransferRecord.from_processor_id == from_pid,
                TransferRecord.to_processor_id == to_pid,
                TransferRecord.quantity == qty,
                TransferRecord.transfer_date == transfer_date
            ).first()
            if existing2:
                transfer_seen.add(key)
                return False
        transfer_seen.add(key)
        return True

    for fpath in xlsx_files:
        fname = os.path.basename(fpath)
        processor_name = extract_processor_name(fname)

        file_report = {'file': fname, 'processor': processor_name,
                       'sheets': [], 'products_found': 0,
                       'shipments': 0, 'receipts': 0, 'begin_count': 0}

        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            report['errors'].append(f'{fname}: 无法打开 - {str(e)}')
            continue

        # 确保加工户存在
        processor = Processor.query.filter_by(name=processor_name).first()
        if not processor:
            processor = Processor(name=processor_name, notes='从Excel导入')
            db.session.add(processor)
            db.session.commit()

        # --- 解析 Sheet: 发货 ---
        if '发货' in wb.sheetnames:
            ws = wb['发货']
            headers = [cell.value for cell in ws[1]]
            # 尝试找日期列和产品列
            date_col = name_col = qty_col = spec_col = remark_col = None
            for idx, h in enumerate(headers):
                if h and '日期' in str(h):
                    date_col = idx
                elif h and ('半成品' in str(h) or '名称' in str(h)):
                    name_col = idx
                elif h and ('发出数量' in str(h) or '数量' in str(h)):
                    qty_col = idx
                elif h and ('备注' in str(h)):
                    remark_col = idx  # 备注列，用于转单检测

            # 检测颜色/规格列（产品名后面的无名列，优先级高于备注列）
            color_col = None
            if name_col is not None:
                cc = name_col + 1
                if cc < len(headers) and (headers[cc] is None or str(headers[cc]).strip() == ''):
                    color_samples = set()
                    for row in ws.iter_rows(min_row=2, max_row=min(15, ws.max_row), values_only=True):
                        vals = list(row)
                        if cc < len(vals) and vals[cc] is not None:
                            v = str(vals[cc]).strip()
                            if v and v != 'None' and not v.replace('.','').replace('-','').isdigit():
                                color_samples.add(v)
                    if color_samples:
                        color_col = cc

            if name_col is not None and qty_col is not None:
                current_date = ''
                shipment = None
                row_count = 0
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    vals = list(row)
                    if all(v is None for v in vals):
                        continue
                    cell_date = str(vals[date_col]).strip() if date_col is not None and vals[date_col] else ''
                    cell_name = str(vals[name_col]).strip() if name_col is not None and vals[name_col] else ''
                    cell_qty = vals[qty_col] if qty_col is not None else None
                    cell_color = str(vals[color_col]).strip() if color_col is not None and vals[color_col] else ''
                    cell_remark = str(vals[remark_col]).strip() if remark_col is not None and vals[remark_col] else ''
                    cell_spec = cell_color if cell_color else cell_remark

                    if cell_name and cell_name not in ['None', '']:
                        # 规范化产品名
                        prod_name = normalize_product_name(cell_name)
                        prod = find_or_create_product(prod_name)
                        if prod:
                            qty = parse_float(cell_qty)
                            if qty and qty != 0:
                                # 按日期分组创建一个发货单
                                if cell_date and cell_date != current_date and cell_date != 'None':
                                    current_date = cell_date
                                    ship_date = normalize_date(cell_date)
                                    # 去重：检查是否已导入
                                    existing = Shipment.query.filter_by(
                                        processor_id=processor.id,
                                        shipment_date=ship_date
                                    ).first()
                                    if existing:
                                        shipment = existing
                                    else:
                                        shipment = Shipment(
                                            processor_id=processor.id,
                                            shipment_date=ship_date,
                                            notes=f'从{fname}导入'
                                        )
                                        db.session.add(shipment)
                                        db.session.flush()
                                        row_count += 1

                                if shipment:
                                    item = ShipmentItem(
                                        shipment_id=shipment.id,
                                        product_id=prod.id,
                                        quantity=qty,
                                        spec=cell_spec if cell_spec and cell_spec != 'None' else ''
                                    )
                                    db.session.add(item)

                                    # 检测备注中的转入信息（如 "李春花转过来"）
                                    transfer_text = cell_remark if cell_remark else cell_spec
                                    if transfer_text and '转' in transfer_text:
                                        source_name = extract_transfer_target(transfer_text)
                                        if source_name and source_name != processor.name:
                                            source = Processor.query.filter_by(name=source_name).first()
                                            if source:
                                                transfer_date = normalize_date(cell_date) if cell_date else ship_date
                                                if add_transfer(source.id, processor.id, prod.id, qty, transfer_date):
                                                    tr = TransferRecord(
                                                        from_processor_id=source.id,
                                                        to_processor_id=processor.id,
                                                        product_id=prod.id,
                                                        quantity=qty,
                                                        transfer_date=transfer_date,
                                                        notes=f'从{fname}导入: {transfer_text}'
                                                    )
                                                    db.session.add(tr)

                file_report['shipments'] = row_count
                if row_count > 0:
                    db.session.commit()

        # --- 解析 Sheet: 接货 ---
        if '接货' in wb.sheetnames:
            ws = wb['接货']
            headers = [cell.value for cell in ws[1]]
            date_col = name_col = qty_col = price_col = amount_col = spec_col = remark_col = color_col = None
            for idx, h in enumerate(headers):
                if h:
                    h_str = str(h)
                    if '日期' in h_str:
                        date_col = idx
                    elif '成品' in h_str or '名称' in h_str:
                        name_col = idx
                    elif '入库数量' in h_str or '数量' in h_str:
                        qty_col = idx
                    elif '单价' in h_str:
                        price_col = idx
                    elif '金额' in h_str and '总' not in h_str:
                        amount_col = idx
                    elif '备注' in h_str:
                        remark_col = idx

            # 如果没找到产品名列，保守跳过（但先尝试列1）
            if name_col is None:
                first_data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
                if first_data_row and len(first_data_row) > 1 and first_data_row[1] is not None:
                    name_col = 1
                    if date_col is None:
                        date_col = 0

            # 检测颜色/规格列（产品名后面的无名列）
            if name_col is not None:
                cc = name_col + 1
                if cc < len(headers) and (headers[cc] is None or str(headers[cc]).strip() == ''):
                    color_samples = set()
                    for row in ws.iter_rows(min_row=2, max_row=min(15, ws.max_row), values_only=True):
                        vals = list(row)
                        if cc < len(vals) and vals[cc] is not None:
                            v = str(vals[cc]).strip()
                            if v and v != 'None' and not v.replace('.','').replace('-','').replace(',','').isdigit():
                                color_samples.add(v)
                    if color_samples:
                        color_col = cc

            if name_col is not None and qty_col is not None:
                current_date = ''
                receipt = None
                row_count = 0
                seen_receipts = {}  # recv_date -> receipt object
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    vals = list(row)
                    if all(v is None for v in vals):
                        continue
                    cell_date = str(vals[date_col]).strip() if date_col is not None and vals[date_col] else ''
                    cell_name = str(vals[name_col]).strip() if name_col is not None and vals[name_col] else ''
                    cell_qty = vals[qty_col] if qty_col is not None else None
                    cell_price = vals[price_col] if price_col is not None else None
                    cell_amount = vals[amount_col] if amount_col is not None else None
                    cell_color = str(vals[color_col]).strip() if color_col is not None and vals[color_col] else ''
                    cell_remark = str(vals[remark_col]).strip() if remark_col is not None and vals[remark_col] else ''
                    cell_spec = cell_color if cell_color else cell_remark

                    if cell_name and cell_name not in ['None', '']:
                        prod_name = normalize_product_name(cell_name)
                        prod = find_or_create_product(prod_name)
                        if prod:
                            qty = parse_float(cell_qty)
                            if qty and qty != 0:
                                if cell_date and cell_date != current_date and cell_date != 'None':
                                    current_date = cell_date
                                    recv_date = normalize_date(cell_date)
                                    if recv_date not in seen_receipts:
                                        # 检查数据库中是否已有（跨次导入去重）
                                        existing_rc = Receipt.query.filter_by(
                                            processor_id=processor.id,
                                            receipt_date=recv_date
                                        ).first()
                                        if existing_rc:
                                            receipt = existing_rc
                                        else:
                                            receipt = Receipt(
                                                processor_id=processor.id,
                                                receipt_date=recv_date,
                                                notes=f'从{fname}导入'
                                            )
                                            db.session.add(receipt)
                                            db.session.flush()
                                            row_count += 1
                                        seen_receipts[recv_date] = receipt
                                    else:
                                        receipt = seen_receipts[recv_date]

                                if receipt:
                                    up = parse_float(cell_price) if cell_price else 0
                                    amt = parse_float(cell_amount) if cell_amount else qty * up
                                    item = ReceiptItem(
                                        receipt_id=receipt.id,
                                        product_id=prod.id,
                                        quantity=qty,
                                        unit_price=up,
                                        amount=amt,
                                        spec=cell_spec if cell_spec and cell_spec != 'None' else ''
                                    )
                                    db.session.add(item)
                                    receipt.total_amount = (receipt.total_amount or 0) + amt

                                    # 检测备注中的转单信息
                                    if cell_spec and '转' in cell_spec:
                                        transfer_text = cell_remark if cell_remark else cell_spec
                                        target_name = extract_transfer_target(transfer_text)
                                        if target_name:
                                            target = Processor.query.filter_by(name=target_name).first()
                                            if target and target.id != processor.id:
                                                transfer_date = normalize_date(cell_date) if cell_date else recv_date
                                                if add_transfer(processor.id, target.id, prod.id, qty, transfer_date):
                                                    tr = TransferRecord(
                                                        from_processor_id=processor.id,
                                                        to_processor_id=target.id,
                                                        product_id=prod.id,
                                                        quantity=qty,
                                                        transfer_date=transfer_date,
                                                        notes=f'从{fname}导入: {transfer_text}'
                                                    )
                                                    db.session.add(tr)

                file_report['receipts'] = row_count
                if row_count > 0:
                    db.session.commit()

        # --- 解析 Sheet1: 期初库存 ---
        if 'Sheet1' in wb.sheetnames:
            ws = wb['Sheet1']
            hdr = [cell.value for cell in ws[1]]
            prod_col = None
            qty_col = None
            for idx, h in enumerate(hdr):
                if h:
                    hs = str(h)
                    if '名称' in hs or '物料' in hs:
                        prod_col = idx
                    elif '期初' in hs:
                        qty_col = idx
            if prod_col is None:
                prod_col = 0
            if qty_col is None:
                qty_col = 1

            begin_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                vals = list(row)
                if prod_col is None or prod_col >= len(vals):
                    continue
                cell_name = str(vals[prod_col]).strip() if vals[prod_col] else ''
                if not cell_name or cell_name == 'None':
                    continue
                try:
                    qty = float(str(vals[qty_col]).strip()) if vals[qty_col] else 0
                except (ValueError, TypeError):
                    qty = 0
                if qty == 0:
                    continue

                prod_name = normalize_product_name(cell_name)
                prod = find_or_create_product(prod_name)
                if prod:
                    existing = BeginInventory.query.filter_by(
                        product_id=prod.id,
                        processor_id=processor.id
                    ).first()
                    if existing:
                        existing.quantity = qty
                    else:
                        bi = BeginInventory(
                            product_id=prod.id,
                            processor_id=processor.id,
                            quantity=qty,
                            period='2026-02'
                        )
                        db.session.add(bi)
                        begin_count += 1
            if begin_count > 0:
                db.session.commit()
                file_report['begin_count'] = begin_count

        report['files'].append(file_report)
        report['total_products'] += file_report['products_found']
        report['total_shipments'] += file_report['shipments']
        report['total_receipts'] += file_report['receipts']

    # 数据体检
    anomalies = run_data_check()

    return render_template('import_report.html', report=report,
                           anomalies=anomalies)


def normalize_product_name(name):
    """规范化产品名称，去除多余空格等"""
    name = name.strip()
    # 统一括号
    name = name.replace('（', '(').replace('）', ')')
    return name


def find_or_create_product(name):
    """查找或创建产品"""
    if not name or name == 'None':
        return None
    # 先精确查找
    prod = Product.query.filter_by(name=name).first()
    if prod:
        return prod
    # 不存在则创建
    prod = Product(name=name, unit='个')
    db.session.add(prod)
    db.session.flush()
    return prod


def parse_float(val):
    """安全解析浮点数"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0


def normalize_date(val):
    """规范化日期"""
    if not val or val == 'None':
        return datetime.now().strftime('%Y-%m-%d')
    # Excel 序列号
    try:
        serial = float(val)
        if serial > 40000:
            # Excel 日期序列号从1900-01-01开始
            base = datetime(1899, 12, 30)
            d = base + timedelta(days=int(serial))
            return d.strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        pass

    s = str(val).strip()
    # 处理 "2.6" -> "2026-02-06" 或 "2.6"
    parts = s.replace('/', '.').split('.')
    if len(parts) == 2:
        m, d = parts
        return f'2026-{int(m):02d}-{int(d):02d}'
    if len(parts) == 3:
        _, m, d = parts
        return f'2026-{int(m):02d}-{int(d):02d}'
    return s


def run_data_check():
    """数据体检 - 检查异常"""
    anomalies = []
    inv_data = get_inventory()

    # 累计发货和收货
    for row in inv_data:
        if row.balance < -0.01:
            anomalies.append({
                'type': 'negative_balance',
                'processor': row.processor_name,
                'product': row.product_name,
                'begin_qty': row.begin_qty,
                'ship_qty': row.ship_qty,
                'receipt_qty': row.receipt_qty,
                'balance': row.balance,
                'msg': f'{row.processor_name} - {row.product_name}: 结存={row.balance:.1f}（负数）'
            })
        elif row.ship_qty == 0 and row.transfer_in_qty == 0 and row.receipt_qty > 0:
            anomalies.append({
                'type': 'receipt_no_shipment',
                'processor': row.processor_name,
                'product': row.product_name,
                'msg': f'{row.processor_name} - {row.product_name}: 有收货({row.receipt_qty:.0f})但无发货或转入记录'
            })

    # 检查单价为0的收货（批量关联加载避免 N+1）
    zero_price_items = db.session.query(ReceiptItem).options(
        joinedload(ReceiptItem.product),
        joinedload(ReceiptItem.receipt).joinedload(Receipt.processor)
    ).filter(ReceiptItem.unit_price == 0, ReceiptItem.quantity > 0).limit(20).all()
    for item in zero_price_items:
        p = item.receipt.processor if item.receipt else None
        prod = item.product
        anomalies.append({
            'type': 'zero_price',
            'processor': p.name if p else '?',
            'product': prod.name if prod else '?',
            'msg': f'{p.name if p else "?"} - {prod.name if prod else "?"}: 收货单价为0'
        })

    return anomalies


# ---------------------------------------------------------------------------
# 路由 - API（供前端调用）
# ---------------------------------------------------------------------------

@app.route('/api/products')
def api_products():
    products = Product.query.order_by(Product.name).all()
    return jsonify([p.to_dict() for p in products])


@app.route('/api/processors')
def api_processors():
    processors = Processor.query.order_by(Processor.name).all()
    return jsonify([p.to_dict() for p in processors])


@app.route('/api/processor/<int:pid>/products')
def api_processor_products(pid):
    """返回全部产品及该加工户的定价（无定价则为0）"""
    products = Product.query.order_by(Product.name).all()
    price_map = {}
    for pp in ProcessorPrice.query.filter_by(processor_id=pid).all():
        price_map[pp.product_id] = pp.price or 0
    result = []
    for prod in products:
        result.append({
            'id': prod.id,
            'name': prod.name,
            'unit': prod.unit or '个',
            'product_type': prod.product_type or '成品',
            'price': price_map.get(prod.id, 0),
        })
    return jsonify(result)


@app.route('/api/products/<int:id>/rename', methods=['POST'])
def api_product_rename(id):
    """修改产品名称"""
    prod = Product.query.get_or_404(id)
    data = request.get_json()
    new_name = data.get('name', '').strip()
    if not new_name:
        return jsonify({'ok': False, 'msg': '产品名称不能为空'})
    prod.name = new_name
    db.session.commit()
    return jsonify({'ok': True, 'name': new_name})


# ---------------------------------------------------------------------------
# 启动
# ---------------------------------------------------------------------------

def init_db():
    """初始化数据库并创建默认数据"""
    with app.app_context():
        db.create_all()

        # 兼容旧数据库：加工户增加工序字段
        try:
            db.session.execute(text(
                "ALTER TABLE processors ADD COLUMN process_step VARCHAR(50) DEFAULT ''"))
            db.session.commit()
        except Exception:
            pass  # 字段已存在

        # 移除 products.name 的唯一约束（SQLite 需重建表）
        try:
            idx = db.session.execute(text(
                "SELECT name FROM sqlite_master WHERE type='index' AND name='sqlite_autoindex_products_1'"
            )).fetchone()
            if idx:
                db.session.execute(text("PRAGMA foreign_keys=OFF"))
                db.session.execute(text("""
                    CREATE TABLE products_new (
                        id INTEGER NOT NULL PRIMARY KEY,
                        name VARCHAR(200) NOT NULL,
                        unit VARCHAR(20) DEFAULT '个',
                        category_id INTEGER,
                        product_type VARCHAR(20) DEFAULT '成品',
                        notes TEXT,
                        created_at DATETIME
                    )
                """))
                db.session.execute(text("INSERT INTO products_new SELECT * FROM products"))
                db.session.execute(text("DROP TABLE products"))
                db.session.execute(text("ALTER TABLE products_new RENAME TO products"))
                db.session.commit()
        except Exception:
            db.session.rollback()
        finally:
            db.session.execute(text("PRAGMA foreign_keys=ON"))

        # 创建默认分类
        if ProductCategory.query.count() == 0:
            for cat in ['毛绒玩具', '服装', '配件', '布匹']:
                db.session.add(ProductCategory(name=cat))
            db.session.commit()


def run_server():
    init_db()
    print('=' * 60)
    print('  委托加工管理系统已启动')
    print('  ' + '=' * 60)
    print(f'  访问地址: http://127.0.0.1:5000')
    print(f'  按 Ctrl+C 停止服务器')
    print('=' * 60)
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', '1') == '1'
    app.run(host='0.0.0.0', port=port, debug=debug)


if __name__ == '__main__':
    run_server()
