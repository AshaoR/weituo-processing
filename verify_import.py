#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""核验导入数据 vs 原始 Excel — 按产品名汇总数量对比"""

import openpyxl, os, sqlite3, sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')
conn = sqlite3.connect(r'C:\Users\邵长荣\委托加工系统\data\database.db')
c = conn.cursor()

def nname(name):
    return name.strip().replace('（', '(').replace('）', ')')

def pfloat(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').strip())
    except:
        return 0

def ndate(val):
    if not val:
        return None
    from datetime import datetime, timedelta
    s = str(val).strip()
    try:
        serial = float(s)
        if serial > 40000:
            base = datetime(1899, 12, 30)
            d = base + timedelta(days=int(serial))
            return d.strftime('%Y-%m-%d')
    except:
        pass
    parts = s.replace('/', '.').split('.')
    if len(parts) == 2:
        m, d = parts
        return f'2026-{int(m):02d}-{int(d):02d}'
    if len(parts) == 3:
        y, m, d = parts
        if len(y) == 2: y = '20' + y
        return f'{y}-{int(m):02d}-{int(d):02d}'
    return s

c.execute('SELECT id, name FROM processors ORDER BY name')
processors = {}
for row in c.fetchall():
    processors[row[1]] = row[0]

data_dir = r'D:\wechat\jilu\xwechat_files\wxid_ufogibn91jth22_59b4\msg\file\2026-05\委托加工物资'

total_err = 0

for fname in sorted(os.listdir(data_dir)):
    if not fname.endswith('.xlsx') or fname.startswith('~$'):
        continue
    pname = fname.replace('.xlsx', '')
    if pname not in processors:
        print(f'[SKIP] {pname}: 不在数据库中')
        continue
    pid = processors[pname]
    fpath = os.path.join(data_dir, fname)
    wb = openpyxl.load_workbook(fpath, data_only=True)

    # ===== SHIPMENTS =====
    if '发货' in wb.sheetnames:
        ws = wb['发货']
        hdr = [cell.value for cell in ws[1]]
        dc = nc = qc = None
        for i, h in enumerate(hdr):
            if h:
                if '日期' in str(h): dc = i
                elif '半成品' in str(h) or '名称' in str(h): nc = i
                elif '发出数量' in str(h) or '数量' in str(h): qc = i
        if nc is not None and qc is not None:
            ex = defaultdict(lambda: defaultdict(float))
            cur_d = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                vals = list(row)
                if all(v is None for v in vals):
                    continue
                cd = str(vals[dc]).strip() if dc is not None and vals[dc] else ''
                cn = str(vals[nc]).strip() if nc is not None and vals[nc] else ''
                cq = vals[qc] if qc is not None else None
                if cn and cn != 'None':
                    pn = nname(cn)
                    q = pfloat(cq)
                    # Check if this column actually contains quantities (not color names)
                    if q and q != 0:
                        if cd and cd != cur_d and cd != 'None':
                            cur_d = cd
                        if cur_d:
                            nd = ndate(cur_d)
                            ex[nd][pn] += q
            # DB
            c.execute('SELECT id, shipment_date FROM shipments WHERE processor_id=?', (pid,))
            db = {}
            for sid, sd in c.fetchall():
                db[sd] = sid
            for d in sorted(set(ex.keys()) | set(db.keys())):
                ei = ex.get(d, {})
                di = {}
                if d in db:
                    c.execute('''SELECT p.name, SUM(si.quantity)
                        FROM shipment_items si JOIN products p ON p.id=si.product_id
                        WHERE si.shipment_id=? GROUP BY p.id''', (db[d],))
                    for pn, q in c.fetchall():
                        di[pn] = q
                for pn in sorted(set(ei.keys()) | set(di.keys())):
                    eq = ei.get(pn, 0)
                    dq = di.get(pn, 0)
                    if abs(eq - dq) > 0.01:
                        print(f'[SHIP] {pname} {d}: [{pn}] Excel={eq:.1f} DB={dq:.1f}')
                        total_err += 1

    # ===== RECEIPTS =====
    if '接货' in wb.sheetnames:
        ws = wb['接货']
        hdr = [cell.value for cell in ws[1]]
        dc = nc = qc = pc = ac = None
        for i, h in enumerate(hdr):
            if h:
                hs = str(h)
                if '日期' in hs: dc = i
                elif '成品' in hs or '名称' in hs: nc = i
                elif '入库数量' in hs or '数量' in hs: qc = i
                elif '单价' in hs: pc = i
                elif '金额' in hs and '总' not in hs: ac = i
        if nc is None: nc = 1
        if dc is None: dc = 0
        if qc is not None:
            ex_q = defaultdict(lambda: defaultdict(float))
            ex_a = defaultdict(lambda: defaultdict(float))
            cur_d = None
            for row in ws.iter_rows(min_row=2, values_only=True):
                vals = list(row)
                if all(v is None for v in vals):
                    continue
                cd = str(vals[dc]).strip() if dc is not None and vals[dc] else ''
                cn = str(vals[nc]).strip() if nc is not None and vals[nc] else ''
                cq = vals[qc] if qc is not None else None
                if cn and cn != 'None':
                    pn = nname(cn)
                    q = pfloat(cq)
                    if q and q != 0:
                        if cd and cd != cur_d and cd != 'None':
                            cur_d = cd
                        if cur_d:
                            nd = ndate(cur_d)
                            ex_q[nd][pn] += q
                            up = pfloat(vals[pc]) if pc is not None and vals[pc] else 0
                            amt = pfloat(vals[ac]) if ac is not None and vals[ac] else 0
                            ex_a[nd][pn] += amt if amt else q * up
            # DB
            c.execute('SELECT id, receipt_date FROM receipts WHERE processor_id=?', (pid,))
            db = {}
            for rid, rd in c.fetchall():
                db[rd] = rid
            for d in sorted(set(ex_q.keys()) | set(db.keys())):
                ei_q = ex_q.get(d, {})
                ei_a = ex_a.get(d, {})
                di = {}
                if d in db:
                    c.execute('''SELECT p.name, SUM(ri.quantity), SUM(ri.amount)
                        FROM receipt_items ri JOIN products p ON p.id=ri.product_id
                        WHERE ri.receipt_id=? GROUP BY p.id''', (db[d],))
                    for pn, q, a in c.fetchall():
                        di[pn] = (q, a)
                for pn in sorted(set(ei_q.keys()) | set(di.keys())):
                    eq = ei_q.get(pn, 0)
                    dq, da = di.get(pn, (0, 0))
                    if abs(eq - dq) > 0.01:
                        print(f'[RECV] {pname} {d}: [{pn}] 数量 Excel={eq:.1f} DB={dq:.1f}')
                        total_err += 1
                    ea = ei_a.get(pn, 0)
                    if abs(ea - da) > 0.01:
                        print(f'[RECV] {pname} {d}: [{pn}] 金额 Excel={ea:.2f} DB={da:.2f}')
                        total_err += 1

    wb.close()

print(f'\n总计差异: {total_err} 处')
if total_err == 0:
    print('所有数据核验通过！')
else:
    print(f'发现 {total_err} 处差异，请检查')
